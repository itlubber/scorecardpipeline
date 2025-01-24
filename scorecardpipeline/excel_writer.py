# -*- coding: utf-8 -*-
"""
@Time    : 2023/05/14 16:23
@Author  : itlubber
@Site    : itlubber.art
"""
import sys
import warnings

warnings.filterwarnings("ignore")

import re
import os
import copy
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from openpyxl.cell.cell import Cell
from openpyxl.drawing.image import Image
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import NamedStyle, Border, Side, Alignment, PatternFill, Font


class ExcelWriter:

    def __init__(self, style_excel=None, style_sheet_name="初始化", mode="replace", fontsize=10, font='楷体', theme_color='2639E9', opacity=0.85, system=None):
        """excel 写入方法

        :param style_excel: 样式模版文件，默认安装包路径下的 template.xlsx ，如果路径调整需要进行相应的调整
        :param style_sheet_name: 模版文件内初始样式sheet名称，默认即可
        :param mode: 写入模式，默认 replace，可选 replace、append，当选择 append 模式时会将已存在的excel中的内容复制到新的文件中
        :param fontsize: 插入excel文件中内容的字体大小，默认 10
        :param font: 插入excel文件中内容的字体，默认 楷体
        :param theme_color: 主题色，默认 2639E9，注意不包含 #
        :param system: excel报告适配的系统，默认 mac，可选 windows、linux，设置为 windows 时会重新适配 picture 的大小
        :param opacity: 写入dataframe时使用颜色填充主题色的透明度设置，默认 0.85
        """
        self.system = system

        if self.system is None:
            self.system = "mac" if sys.platform == "darwin" else "windows"

        self.english_width = 0.12
        self.chinese_width = 0.21
        self.mode = mode
        self.font = font
        self.opacity = opacity
        self.fontsize = fontsize
        self.theme_color = theme_color
        self.workbook = load_workbook(style_excel or os.path.join(os.path.dirname(os.path.abspath(__file__)), 'template.xlsx'))
        self.style_sheet = self.workbook[style_sheet_name]

        self.name_styles = []
        self.init_style(font, fontsize, theme_color)
        for style in self.name_styles:
            if style.name not in self.workbook.style_names:
                self.workbook.add_named_style(style)

    def add_conditional_formatting(self, worksheet, start_space, end_space):
        """
        设置条件格式

        :param worksheet: 当前选择设置条件格式的sheet
        :param start_space: 开始单元格位置
        :param end_space: 结束单元格位置
        """
        worksheet.conditional_formatting.add(f'{start_space}:{end_space}', DataBarRule(start_type='min', end_type='max', color=self.theme_color))

    @staticmethod
    def set_column_width(worksheet, column, width):
        """
        调整excel列宽

        :param worksheet: 当前选择调整列宽的sheet
        :param column: 列，可以直接输入 index 或者 字母
        :param width: 设置列的宽度
        """
        worksheet.column_dimensions[column if isinstance(column, str) else get_column_letter(column)] = width

    @staticmethod
    def set_number_format(worksheet, space, _format):
        """
        设置数值显示格式

        :param worksheet: 当前选择调整数值显示格式的sheet
        :param space: 单元格范围
        :param _format: 显示格式，参考 openpyxl
        """
        cells = worksheet[space]
        if isinstance(cells, Cell):
            cells = [cells]

        for cell in cells:
            if isinstance(cell, tuple):
                for c in cell:
                    c.number_format = _format
            else:
                cell.number_format = _format

    def set_freeze_panes(self, worksheet, space):
        """
        设置数值显示格式

        :param worksheet: 当前选择调整数值显示格式的sheet
        :param space: 单元格范围
        """
        if not isinstance(worksheet, Worksheet):
            worksheet = self.get_sheet_by_name(worksheet)

        if isinstance(space, (tuple, list)):
            space = self.get_cell_space(space)

        worksheet.freeze_panes = space

    def get_sheet_by_name(self, name):
        """
        获取sheet名称为name的工作簿，如果不存在，则从初始模版文件中拷贝一个名称为name的sheet

        :param name: 需要获取的工作簿名称
        """
        if name not in self.workbook.sheetnames:
            worksheet = self.workbook.copy_worksheet(self.style_sheet)
            worksheet.title = name
        else:
            worksheet = self.workbook[name]

        return worksheet

    def move_sheet(self, worksheet, offset: int = 0, index: int = None):
        """移动 sheet 位置

        :param worksheet: 需要移动的sheet，支持输入字符串和Worksheet
        :param offset: 需要移动的相对位置，默认 0，在传入 index 时参数不生效
        :param index: 需要移动到的位置索引，超出移动到最后
        """
        total_sheets = len(self.workbook.sheetnames)
        if index:
            offset = -(total_sheets - 1) + index

            if offset >= total_sheets:
                offset = total_sheets - 1

        self.workbook.move_sheet(worksheet, offset=offset)

    def insert_hyperlink2sheet(self, worksheet, insert_space, hyperlink=None, file=None, sheet=None, target_space=None):
        """向sheet中的某个单元格插入超链接

        :param worksheet: 需要插入超链接的sheet
        :param insert_space: 超链接插入的单元格位置，可以是 "B2" 或者 (2, 2) 任意一种形式，首个单元格从 (1, 1) 开始
        :param hyperlink: 超链接的地址, 与 target_space 参数互斥，优先 hyperlink，格式参考: [f"file://{文件地址} - #{表名}!{单元格位置}", f"#{表名}!{单元格位置}", f"{单元格位置}"], 其中 单元格位置 为类似 "B2" 的格式
        :param file: 超链接的文件路径，默认 None，即当前excel文件，传入 hyperlink 参数时无效，传入file时确保sheet参数已传，否则默认为当前sheet
        :param sheet: 超链接的sheet名称，默认 None，即当前sheet，传入 hyperlink 参数时无效
        :param target_space: 超链接的单元格位置，默认 None，支持 "B2" 或者 (2, 2) 任意一种形式，传入 hyperlink 参数时无效
        """
        if isinstance(insert_space, str):
            start_col = re.findall('\D+', insert_space)[0]
            start_row = int(re.findall("\d+", insert_space)[0])
        else:
            start_col = get_column_letter(insert_space[1])
            start_row = insert_space[0]

        cell = worksheet[f"{start_col}{start_row}"]

        if hyperlink is None:
            if target_space is None:
                raise ValueError("hyperlink 和 target_space 二选一，必须选择传入一个")

            if sheet is None:
                sheet = worksheet.title

            if isinstance(target_space, str):
                target_col = re.findall('\D+', target_space)[0]
                target_row = int(re.findall("\d+", target_space)[0])
            else:
                target_col = get_column_letter(target_space[1])
                target_row = target_space[0]

            if file:
                hyperlink = f"file://{file} - #{sheet}!{target_col}{target_row}"
            else:
                hyperlink = f"#{sheet}!{target_col}{target_row}"

        cell.hyperlink = Hyperlink(ref=f"{start_col}{start_row}", location=hyperlink, display=f"{cell.value}")

    def insert_value2sheet(self, worksheet, insert_space, value="", style="content", auto_width=False, end_space=None, align: dict=None, max_col_width=50):
        """
        向sheet中的某个单元格插入某种样式的内容

        :param worksheet: 需要插入内容的sheet
        :param insert_space: 内容插入的单元格位置，可以是 "B2" 或者 (2, 2) 任意一种形式
        :param value: 需要插入的内容
        :param style: 渲染的样式，参考 init_style 中初始设置的样式
        :param end_space: 如果需要合并单元格，传入需要截止的单元格位置信息，可以是 "B2" 或者 (2, 2) 任意一种形式
        :param auto_width: 是否开启自动调整列宽
        :param align: 文本排列方式, 参考: Alignment
        :param max_col_width: 单元格列最大宽度，默认 50

        :return: 返回插入元素最后一列之后、最后一行之后的位置
        """
        if isinstance(insert_space, str):
            start_col = re.findall('\D+', insert_space)[0]
            start_row = int(re.findall("\d+", insert_space)[0])
        else:
            start_col = get_column_letter(insert_space[1])
            start_row = insert_space[0]

        cell = worksheet[f"{start_col}{start_row}"]
        cell.style = style

        if align:
            _align = {"horizontal": "center", "vertical": "center"}
            _align.update(align)
            cell.alignment = Alignment(**_align)

        if end_space is not None:
            if isinstance(end_space, str):
                end_col = re.findall('\D+', end_space)[0]
                end_row = int(re.findall("\d+", end_space)[0])
            else:
                end_col = get_column_letter(end_space[1])
                end_row = end_space[0]

            worksheet.merge_cells(f"{start_col}{start_row}:{end_col}{end_row}")

        worksheet[f"{start_col}{start_row}"] = value

        if auto_width:
            # original_styles = [worksheet[f"{start_col}{i}"].fill.copy() for i in range(1, worksheet.max_row + 1)]
            curr_width = worksheet.column_dimensions[start_col].width
            auto_width = min(max([(self.check_contain_chinese(value)[1] * self.english_width + self.check_contain_chinese(value)[2] * self.chinese_width) * self.fontsize, 10, curr_width]), max_col_width)
            worksheet.column_dimensions[start_col].width = auto_width

            # for i in range(worksheet.max_row):
            #     worksheet[f"{start_col}{i + 1}"].fill = original_styles[i]

        if end_space is not None:
            return end_row + 1, column_index_from_string(end_col) + 1
        else:
            return start_row + 1, column_index_from_string(start_col) + 1

    def insert_pic2sheet(self, worksheet, fig, insert_space, figsize=(600, 250)):
        """
        向excel中插入图片内容

        :param worksheet: 需要插入内容的sheet
        :param fig: 需要插入的图片路径
        :param insert_space: 插入图片的起始单元格
        :param figsize: 图片大小设置
        :return: 返回插入元素最后一列之后、最后一行之后的位置
        """
        if isinstance(insert_space, str):
            start_row = int(re.findall("\d+", insert_space)[0])
            start_col = re.findall('\D+', insert_space)[0]
        else:
            start_row, start_col = insert_space
            start_col = get_column_letter(start_col)

        image = Image(fig)
        image.width, image.height = figsize
        worksheet.add_image(image, f"{start_col}{start_row}")

        return start_row + int(figsize[1] / (16.0 if self.system != 'mac' else 17.5)), column_index_from_string(start_col) + 8

    def insert_rows(self, worksheet, row, row_index, col_index, merge_rows=None, style="", auto_width=False, style_only=False, multi_levels=False):
        """
        向excel中插入一行数据，insert_df2sheet 依赖本方法

        :param worksheet: 需要插入内容的sheet
        :param row: 数据内容
        :param row_index: 插入数据的行索引，用来判断使用哪种边框样式
        :param col_index: 插入数据的列索引，用来判断使用哪种边框样式
        :param merge_rows: 需要合并单元的行索引
        :param style: 插入数据的excel风格
        :param auto_width: 是否自动调整列宽，自动调整列宽会导致该列样式模版发生变化，非内容列默认填充的白色失效
        :param style_only: 是否使用填充样式
        :param multi_levels: 是否多层索引或多层级列
        """
        curr_col = column_index_from_string(col_index)

        if multi_levels and style == "header":
            row = pd.Series(row).ffill().to_list()
            item, start, length = self.calc_continuous_cnt(row)
            while start is not None:
                if start + length < len(row):
                    if start == 0:
                        self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + start)}{row_index}', self.astype_insertvalue(item), style=f"{style}_left" if style else "left", auto_width=auto_width, end_space=f'{get_column_letter(curr_col + start + length - 1)}{row_index}')
                    else:
                        self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + start)}{row_index}', self.astype_insertvalue(item), style=f"{style}_middle" if style else "middle", auto_width=auto_width, end_space=f'{get_column_letter(curr_col + start + length - 1)}{row_index}')
                else:
                    self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + start)}{row_index}', self.astype_insertvalue(item), style=f"{style}_right" if style else "right", auto_width=auto_width, end_space=f'{get_column_letter(curr_col + start + length - 1)}{row_index}')

                item, start, length = self.calc_continuous_cnt(row, start + length)
        else:
            for j, v in enumerate(row):
                if merge_rows is not None and row_index + 1 not in merge_rows:
                    if j == 0:
                        self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + j)}{row_index}', self.astype_insertvalue(v), style="merge_left", auto_width=auto_width)
                    elif j == len(row) - 1:
                        self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + j)}{row_index}', self.astype_insertvalue(v), style="merge_right", auto_width=auto_width)
                    else:
                        self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + j)}{row_index}', self.astype_insertvalue(v), style="merge_middle", auto_width=auto_width)
                elif style_only or len(row) <= 1:
                    self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + j)}{row_index}', self.astype_insertvalue(v), style=style or "middle", auto_width=auto_width)
                else:
                    if j == 0:
                        self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + j)}{row_index}', self.astype_insertvalue(v), style=f"{style}_left" if style else "left", auto_width=auto_width)
                    elif j == len(row) - 1:
                        self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + j)}{row_index}', self.astype_insertvalue(v), style=f"{style}_right" if style else "right", auto_width=auto_width)
                    else:
                        self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + j)}{row_index}', self.astype_insertvalue(v), style=f"{style}_middle" if style else "middle", auto_width=auto_width)

    def insert_df2sheet(self, worksheet, data, insert_space, merge_column=None, header=True, index=False, auto_width=False, fill=False, merge=False, merge_index=True):
        """
        向excel文件中插入指定样式的dataframe数据

        :param worksheet: 需要插入内容的sheet
        :param data: 需要插入的dataframe
        :param insert_space: 插入内容的起始单元格位置
        :param merge_column: 需要分组显示的列，index或者列名，需要提前排好序，从 0.1.33 开始 ExcleWriter 不会自动处理顺序
        :param header: 是否存储dataframe的header，暂不支持多级表头，从 0.1.30 开始支持多层表头和多层索引
        :param index: 是否存储dataframe的index
        :param merge_index: 当存储dataframe的index时，index中同一层级连续相同值是否合并，默认 True，即合并
        :param auto_width: 是否自动调整列宽，自动调整列宽会导致该列样式模版发生变化，非内容列默认填充的白色失效
        :param fill: 是否使用颜色填充而非边框，当 fill 为 True 时，merge_column 失效
        :param merge: 是否合并单元格，配合 merge_column 一起使用，当前版本仅在 merge_column 只有一列时有效
        :return: 返回插入元素最后一列之后、最后一行之后的位置
        """
        df = data.copy()

        if isinstance(insert_space, str):
            start_row = int(re.findall("\d+", insert_space)[0])
            start_col = re.findall('\D+', insert_space)[0]
        else:
            start_row, start_col = insert_space
            start_col = get_column_letter(start_col)

        def get_merge_rows(values, start_row):
            _rows = []
            item, start, length = self.calc_continuous_cnt(values)
            while start is not None:
                _rows.append(start + start_row)
                item, start, length = self.calc_continuous_cnt(values, start + length)

            _rows.append(len(values) + start_row)
            return _rows

        if index and merge_index:
            merge_index_cols = {i: get_column_letter(column_index_from_string(start_col) + i) for i in range(df.index.nlevels)}
            merge_index_rows = {i: get_merge_rows(df.index.get_level_values(i).tolist(), start_row + df.columns.nlevels if header else start_row) for i in range(df.index.nlevels)}
        else:
            merge_index_cols = None
            merge_index_rows = None

        if merge_column:
            if not isinstance(merge_column, (list, np.ndarray)):
                merge_column = [merge_column]

            if isinstance(merge_column[0], (int, float)) and (merge_column[0] not in df.columns):
                merge_column = [df.columns.tolist()[col] if col not in df.columns else col for col in merge_column]

            # 从 0.1.33 开始不修改需要保存的原始数据
            # if df[merge_column].values.tolist() != df[merge_column].sort_values(merge_column).values.tolist():
            #     df = df.sort_values(merge_column)

            if index:
                # merge_cols = [get_column_letter(df.columns.get_loc(col) + column_index_from_string(start_col) + df.index.nlevels) for col in merge_column]
                merge_cols = {col: get_column_letter(df.columns.get_loc(col) + column_index_from_string(start_col) + df.index.nlevels) for col in merge_column}
            else:
                # merge_cols = [get_column_letter(df.columns.get_loc(col) + column_index_from_string(start_col)) for col in merge_column]
                merge_cols = {col: get_column_letter(df.columns.get_loc(col) + column_index_from_string(start_col)) for col in merge_column}

            if header:
                # merge_rows = list(np.cumsum(df.groupby(merge_column)[merge_column].count().values[:, 0]) + start_row + df.columns.nlevels)
                merge_rows = {col: get_merge_rows(df[col].tolist(), start_row + df.columns.nlevels) for col in merge_column}
            else:
                # merge_rows = list(np.cumsum(df.groupby(merge_column)[merge_column].count().values[:, 0]) + start_row)
                merge_rows = {col: get_merge_rows(df[col].tolist(), start_row) for col in merge_column}
        else:
            merge_cols = None
            merge_rows = None

        def _iter_rows(df, header=True, index=True):
            columns = df.columns.tolist()
            indexs = df.index.tolist()
            for i, row in enumerate(dataframe_to_rows(df, header=header, index=False)):
                if header:
                    if i < df.columns.nlevels:
                        if index:
                            if df.columns.nlevels > 1:
                                if i == df.columns.nlevels - 1:
                                    yield list(df.index.names) + [c[i] for c in columns]
                                else:
                                    yield [None] * df.index.nlevels + [c[i] for c in columns]
                            else:
                                yield list(df.index.names) + columns
                        else:
                            if df.columns.nlevels > 1 and i < df.columns.nlevels:
                                yield [c[i] for c in columns]
                            else:
                                yield columns
                    else:
                        if index:
                            if df.index.nlevels > 1:
                                yield list(indexs[int(i - df.columns.nlevels)]) + row
                            else:
                                yield [indexs[int(i - df.columns.nlevels)]] + row
                        else:
                            yield row
                else:
                    if index:
                        if df.index.nlevels > 1:
                            yield list(indexs[i]) + row
                        else:
                            yield [indexs[i]] + row
                    else:
                        yield row

        for i, row in enumerate(_iter_rows(df, header=header, index=index)):
            if fill:
                if header and i < df.columns.nlevels:
                    self.insert_rows(worksheet, row, start_row + i, start_col, style="header", auto_width=auto_width, multi_levels=True if df.columns.nlevels > 1 else False)
                elif i == 0:
                    self.insert_rows(worksheet, row, start_row + i, start_col, style="middle_even_first", auto_width=auto_width, style_only=True)
                else:
                    if df.columns.nlevels % 2 == 1:
                        if i % 2 == 1:
                            style = "middle_odd_last" if (header and i == len(df) + df.columns.nlevels - 1) or (not header and i + 1 == len(df)) else "middle_odd"
                        else:
                            style = "middle_even_last" if (header and i == len(df) + df.columns.nlevels - 1) or (not header and i + 1 == len(df)) else "middle_even"
                    else:
                        if i % 2 == 1:
                            style = "middle_even_last" if (header and i == len(df) + df.columns.nlevels - 1) or (not header and i + 1 == len(df)) else "middle_even"
                        else:
                            style = "middle_odd_last" if (header and i == len(df) + df.columns.nlevels - 1) or (not header and i + 1 == len(df)) else "middle_odd"
                    self.insert_rows(worksheet, row, start_row + i, start_col, style=style, auto_width=auto_width, style_only=True)
            else:
                if header and i < df.columns.nlevels:
                    self.insert_rows(worksheet, row, start_row + i, start_col, style="header", auto_width=auto_width, multi_levels=True if df.columns.nlevels > 1 else False)
                elif i == 0:
                    self.insert_rows(worksheet, row, start_row + i, start_col, style="first", auto_width=auto_width)
                elif (header and i == len(df) + df.columns.nlevels - 1) or (not header and i + 1 == len(df)):
                    self.insert_rows(worksheet, row, start_row + i, start_col, style="last", auto_width=auto_width)
                else:
                    # self.insert_rows(worksheet, row, start_row + i, start_col, auto_width=auto_width, merge_rows=merge_rows)
                    if merge_rows and len(merge_rows) > 0:
                        self.insert_rows(worksheet, row, start_row + i, start_col, auto_width=auto_width, merge_rows=sorted(set(_row for _rows in merge_rows.values() for _row in _rows)))
                    else:
                        self.insert_rows(worksheet, row, start_row + i, start_col, auto_width=auto_width)

        # 合并索引单元格
        if index and merge_index and merge_index_rows and len(merge_index_rows) > 0:
            for col in merge_index_cols.keys():
                merge_col = merge_index_cols[col]
                merge_row = merge_index_rows[col]

                for s, e in zip(merge_row[:-1], merge_row[1:]):
                    if e - s > 1:
                        self.merge_cells(worksheet, f"{merge_col}{s}", f"{merge_col}{e - 1}")

        # 合并列单元格
        if merge and merge_column and merge_cols and len(merge_cols) > 0:
            for col in merge_cols.keys():
                merge_col = merge_cols[col]
                merge_row = merge_rows[col]

                for s, e in zip(merge_row[:-1], merge_row[1:]):
                    if e - s > 1:
                        self.merge_cells(worksheet, f"{merge_col}{s}", f"{merge_col}{e - 1}")

        end_row = start_row + len(data) + df.columns.nlevels if header else start_row + len(data)

        return end_row, column_index_from_string(start_col) + len(data.columns)

    def merge_cells(self, worksheet, start, end):
        """合并同一列单元格并保证样式相应合并

        :param worksheet: 需要合并单元格的sheet
        :param start: 合并单元格开始的位置
        :param end: 合并单元格结束的位置
        :return:
        """
        if isinstance(start, str):
            start_col, start_row = self.get_cell_space(start)
        elif isinstance(start, (tuple, list)):
            start_col, start_row = start[0], start[1]
        else:
            raise TypeError("仅支持二元组或字符串")

        if isinstance(end, str):
            end_col, end_row = self.get_cell_space(end)
        elif isinstance(end, (tuple, list)):
            end_col, end_row = end[0], end[1]
        else:
            raise TypeError("仅支持二元组或字符串")

        # 判断是否单列合并
        if end_col - start_col != 0:
            raise ValueError("仅支持单列合并单元格")

        # 暂存单元格样式
        cell_style = worksheet[f"{get_column_letter(start_col)}{start_row}"].style
        top_border_style = worksheet[f"{get_column_letter(start_col)}{start_row}"].border.top
        border_style = worksheet[f"{get_column_letter(start_col)}{end_row}"].border
        border_style = Border(
            top=Side(style=top_border_style.style, color=top_border_style.color) if top_border_style else None,
            left=Side(style=border_style.left.style, color=border_style.left.color) if border_style.left else None,
            right=Side(style=border_style.right.style, color=border_style.right.color) if border_style.right else None,
            bottom=Side(style=border_style.bottom.style, color=border_style.bottom.color) if border_style.bottom else None,
        )

        # 将单元格样式应用到需要合并的单元格
        merged_cell = worksheet[f"{get_column_letter(start_col)}{start_row}"]
        merged_cell.style = copy.deepcopy(cell_style)
        merged_cell.border = border_style

        # 合并单元格
        worksheet.merge_cells(f"{get_column_letter(start_col)}{start_row}:{get_column_letter(start_col)}{end_row}")

    @staticmethod
    def check_contain_chinese(check_str):
        """
        检查字符串中是否包含中文

        :param check_str: 需要检查的字符串
        :return: 返回每个字符是否是中文 list<bool>，英文字符个数，中文字符个数
        """
        out = []
        for ch in str(check_str).encode('utf-8').decode('utf-8'):
            if u'\u4e00' <= ch <= u'\u9fff':
                out.append(True)
            else:
                out.append(False)
        return out, len(out) - sum(out), sum(out)

    @staticmethod
    def astype_insertvalue(value, decimal_point=4):
        """
        格式化需要存储excel的内容，如果是浮点型，按照设置的精度进行保存，如果是类别型或其他特殊类型，转字符串存储，如果非以上两种，直接原始值存储

        :param value: 需要插入 excel 的内容
        :param decimal_point: 如果是浮点型，需要保留的精度，默认小数点后4位数
        :return: 格式化后存入excel的内容
        """
        if re.search('tuple|list|set|numpy.ndarray|Categorical|numpy.dtype|Interval', str(type(value))):
            return str(value)
        elif re.search('float', str(type(value))):
            return round(float(value), decimal_point)
        else:
            return value

    @staticmethod
    def calc_continuous_cnt(list_, index_=0):
        """
        根据传入的 list ，计算 list 中某个 index 开始，连续出现该元素的个数

        :param list_: 需要检索的 list
        :param index_: 元素索引

        :return: 元素值，索引值，连续出现的个数

        **参考样例**

        >>> calc_continuous_cnt = ExcelWriter.calc_continuous_cnt
        >>> list_ = ['A','A','A','A','B','C','C','D','D','D']
        >>> calc_continuous_cnt(list_, 0)
        ('A', 0, 4)
        >>> calc_continuous_cnt(list_, 4)
        ('B', 4, 1)
        >>> calc_continuous_cnt(list_, 6)
        ('C', 6, 1)
        """
        if index_ >= len(list_):
            return None, None, None

        else:
            cnt, str_ = 0, list_[index_]
            for i in range(index_, len(list_), 1):
                if list_[i] == str_:
                    cnt = cnt + 1
                else:
                    break
            return str_, index_, cnt

    @staticmethod
    def itlubber_border(border, color, white=False):
        """
        itlubber 的边框样式生成器

        :param border: 边框样式，如果输入长度为 3，则生成 [左，右，下]，如果长度为4，则生成 [左，右，下，上]
        :param color: 边框颜色
        :param white: 是否显示白色边框

        :return: Border
        """
        if len(border) == 3:
            return Border(
                left=Side(border_style=None if not white and color[0] == "FFFFFF" else border[0], color=None if not white and color[0] == "FFFFFF" else color[0]),
                right=Side(border_style=None if not white and color[1] == "FFFFFF" else border[1], color=None if not white and color[1] == "FFFFFF" else color[1]),
                bottom=Side(border_style=border[2], color=color[2]),
            )
        else:
            return Border(
                left=Side(border_style=None if not white and color[0] == "FFFFFF" else border[0], color=None if not white and color[0] == "FFFFFF" else color[0]),
                right=Side(border_style=None if not white and color[1] == "FFFFFF" else border[1], color=None if not white and color[1] == "FFFFFF" else color[1]),
                bottom=Side(border_style=border[2], color=color[2]),
                top=Side(border_style=border[3], color=color[3]),
            )

    @staticmethod
    def get_cell_space(space):
        """
        根据传入的不同格式的位置，转换为另一种形式的excel单元格定位

        :param space: 传入的excel单元格定位，支持两种格式，B1 或 (2, 2)

        :return: 返回单元格定位，tuple / str

        **参考样例**

        >>> get_cell_space = ExcelWriter.get_cell_space
        >>> get_cell_space("B3")
        (2, 3)
        >>> get_cell_space((2, 2))
        'B2'
        """
        if isinstance(space, str):
            start_row = int(re.findall("\d+", space)[0])
            start_col = re.findall('\D+', space)[0]
            return column_index_from_string(start_col), start_row
        else:
            start_row = space[0]
            if isinstance(space[1], int):
                start_col = get_column_letter(space[1])
            else:
                start_col = space[1]
            return f"{start_col}{start_row}"

    @staticmethod
    def calculate_rgba_color(hex_color, opacity, prefix="#"):
        """
        根据某个颜色计算某个透明度对应的颜色

        :param hex_color: hex格式的颜色值
        :param opacity: 透明度，[0, 1] 之间的数值
        :param prefix: 返回颜色的前缀
        :return: 对应某个透明度的颜色
        """
        rgb_color = tuple(int(hex_color.lstrip('#')[i:i + 2], 16) for i in (0, 2, 4))
        rgba_color = tuple(int((1 - opacity) * c + opacity * 255) for c in rgb_color)

        return prefix + '{:02X}{:02X}{:02X}'.format(*rgba_color)

    def init_style(self, font, fontsize, theme_color):
        """
        初始化单元格样式

        :param font: 字体名称
        :param fontsize: 字体大小
        :param theme_color: 主题颜色
        """
        header_style, header_left_style, header_middle_style, header_right_style = NamedStyle(name="header"), NamedStyle(name="header_left"), NamedStyle(name="header_middle"), NamedStyle(name="header_right")
        last_style, last_left_style, last_middle_style, last_right_style = NamedStyle(name="last"), NamedStyle(name="last_left"), NamedStyle(name="last_middle"), NamedStyle(name="last_right")
        content_style, left_style, middle_style, right_style = NamedStyle(name="content"), NamedStyle(name="left"), NamedStyle(name="middle"), NamedStyle(name="right")
        merge_style, merge_left_style, merge_middle_style, merge_right_style = NamedStyle(name="merge"), NamedStyle(name="merge_left"), NamedStyle(name="merge_middle"), NamedStyle(name="merge_right")
        first_style, first_left_style, first_middle_style, first_right_style = NamedStyle(name="first"), NamedStyle(name="first_left"), NamedStyle(name="first_middle"), NamedStyle(name="first_right")

        header_font = Font(size=fontsize, name=font, color="FFFFFF", bold=True)
        header_fill = PatternFill(fill_type="solid", start_color=theme_color)
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        content_fill = PatternFill(fill_type="solid", start_color="FFFFFF")
        content_font = Font(size=fontsize, name=font, color="000000")
        even_fill = PatternFill(fill_type="solid", start_color=self.calculate_rgba_color(self.theme_color, self.opacity, prefix=""))

        header_style.font, header_left_style.font, header_middle_style.font, header_right_style.font = header_font, header_font, header_font, header_font
        header_style.fill, header_left_style.fill, header_middle_style.fill, header_right_style.fill = header_fill, header_fill, header_fill, header_fill
        header_style.alignment, header_left_style.alignment, header_middle_style.alignment, header_right_style.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True), alignment, alignment, alignment

        header_style.border = self.itlubber_border(["medium", "medium", "medium", "medium"], [theme_color, theme_color, theme_color, theme_color], white=True)
        header_left_style.border = self.itlubber_border(["medium", "thin", "medium", "medium"], [theme_color, "FFFFFF", theme_color, theme_color], white=True)
        header_middle_style.border = self.itlubber_border(["thin", "thin", "medium", "medium"], ["FFFFFF", "FFFFFF", theme_color, theme_color], white=True)
        header_right_style.border = self.itlubber_border(["thin", "medium", "medium", "medium"], ["FFFFFF", theme_color, theme_color, theme_color], white=True)

        last_style.font, last_left_style.font, last_middle_style.font, last_right_style.font = content_font, content_font, content_font, content_font
        last_style.fill, last_left_style.fill, last_middle_style.fill, last_right_style.fill = content_fill, content_fill, content_fill, content_fill
        last_style.alignment, last_left_style.alignment, last_middle_style.alignment, last_right_style.alignment = alignment, alignment, alignment, alignment

        last_style.border = self.itlubber_border(["medium", "medium", "medium"], [theme_color, theme_color, theme_color])
        last_left_style.border = self.itlubber_border(["medium", "thin", "medium"], [theme_color, "FFFFFF", theme_color])
        last_middle_style.border = self.itlubber_border(["thin", "thin", "medium"], ["FFFFFF", "FFFFFF", theme_color])
        last_right_style.border = self.itlubber_border(["thin", "medium", "medium"], ["FFFFFF", theme_color, theme_color])

        content_style.font, left_style.font, middle_style.font, right_style.font = content_font, content_font, content_font, content_font
        content_style.fill, left_style.fill, middle_style.fill, right_style.fill = content_fill, content_fill, content_fill, content_fill
        content_style.alignment, left_style.alignment, middle_style.alignment, right_style.alignment = alignment, alignment, alignment, alignment

        content_style.border = self.itlubber_border(["medium", "medium", "thin"], [theme_color, theme_color, theme_color])
        left_style.border = self.itlubber_border(["medium", "thin", "thin"], [theme_color, "FFFFFF", theme_color])
        middle_style.border = self.itlubber_border(["thin", "medium", "thin"], ["FFFFFF", "FFFFFF", theme_color])
        right_style.border = self.itlubber_border(["thin", "medium", "thin"], ["FFFFFF", theme_color, theme_color])

        merge_style.font, merge_left_style.font, merge_middle_style.font, merge_right_style.font = content_font, content_font, content_font, content_font
        merge_style.fill, merge_left_style.fill, merge_middle_style.fill, merge_right_style.fill = content_fill, content_fill, content_fill, content_fill
        merge_style.alignment, merge_left_style.alignment, merge_middle_style.alignment, merge_right_style.alignment = alignment, alignment, alignment, alignment

        merge_style.border = self.itlubber_border(["medium", "medium", "thin"], ["FFFFFF", "FFFFFF", "FFFFFF"])
        merge_left_style.border = self.itlubber_border(["medium", "thin", "thin"], [theme_color, "FFFFFF", "FFFFFF"])
        merge_middle_style.border = self.itlubber_border(["thin", "medium", "thin"], ["FFFFFF", "FFFFFF", "FFFFFF"])
        merge_right_style.border = self.itlubber_border(["thin", "medium", "thin"], ["FFFFFF", theme_color, "FFFFFF"])

        first_style.font, first_left_style.font, first_middle_style.font, first_right_style.font = content_font, content_font, content_font, content_font
        first_style.fill, first_left_style.fill, first_middle_style.fill, first_right_style.fill = content_fill, content_fill, content_fill, content_fill
        first_style.alignment, first_left_style.alignment, first_middle_style.alignment, first_right_style.alignment = alignment, alignment, alignment, alignment

        first_style.border = self.itlubber_border(["medium", "medium", "thin", "medium"], [theme_color, theme_color, theme_color, theme_color])
        first_left_style.border = self.itlubber_border(["medium", "thin", "thin", "medium"], [theme_color, "FFFFFF", theme_color, theme_color])
        first_middle_style.border = self.itlubber_border(["thin", "thin", "thin", "medium"], ["FFFFFF", "FFFFFF", theme_color, theme_color])
        first_right_style.border = self.itlubber_border(["thin", "medium", "thin", "medium"], ["FFFFFF", theme_color, theme_color, theme_color])

        middle_odd_style, middle_odd_first_style, middle_odd_last_style = NamedStyle(name="middle_odd"), NamedStyle(name="middle_odd_first"), NamedStyle(name="middle_odd_last")
        middle_even_style, middle_even_first_style, middle_even_last_style = NamedStyle(name="middle_even"), NamedStyle(name="middle_even_first"), NamedStyle(name="middle_even_last")

        middle_odd_style.font, middle_odd_first_style.font, middle_odd_last_style.font, middle_even_style.font, middle_even_first_style.font, middle_even_last_style.font = content_font, content_font, content_font, content_font, content_font, content_font
        middle_odd_style.alignment, middle_odd_first_style.alignment, middle_odd_last_style.alignment, middle_even_style.alignment, middle_even_first_style.alignment, middle_even_last_style.alignment = alignment, alignment, alignment, alignment, alignment, alignment
        middle_odd_style.fill, middle_odd_first_style.fill, middle_odd_last_style.fill, middle_even_style.fill, middle_even_first_style.fill, middle_even_last_style.fill = content_fill, content_fill, content_fill, even_fill, even_fill, even_fill

        middle_odd_first_style.border = Border(top=Side(border_style="medium", color=self.theme_color))
        middle_odd_last_style.border = Border(bottom=Side(border_style="medium", color=self.theme_color))
        middle_even_first_style.border = Border(top=Side(border_style="medium", color=self.theme_color))
        middle_even_last_style.border = Border(bottom=Side(border_style="medium", color=self.theme_color))
        middle_even_style.border = Border(bottom=Side(border_style="thin", color="FFFFFF"))
        middle_odd_style.border = Border(bottom=Side(border_style="thin", color="FFFFFF"))

        self.name_styles.extend([
            header_style, header_left_style, header_middle_style, header_right_style,
            last_style, last_left_style, last_middle_style, last_right_style,
            content_style, left_style, middle_style, right_style,
            merge_style, merge_left_style, merge_middle_style, merge_right_style,
            first_style, first_left_style, first_middle_style, first_right_style,
            middle_odd_style, middle_even_first_style, middle_odd_last_style, middle_even_style, middle_odd_first_style, middle_even_last_style,
        ])

    def save(self, filename, close=True):
        """
        保存excel文件

        :param filename: 需要保存 excel 文件的路径
        :param close: 是否需要释放 writer
        """
        if self.style_sheet.title in self.workbook.sheetnames:
            self.workbook.remove(self.style_sheet)

        if os.path.exists(filename) and self.mode == "append":
            _workbook = load_workbook(filename)

            for _sheet_name in _workbook.sheetnames:
                if _sheet_name not in self.workbook.sheetnames:
                    _worksheet = self.get_sheet_by_name(_sheet_name)

                    for i, row in enumerate(_workbook[_sheet_name].iter_rows()):
                        for j, cell in enumerate(row):
                            _worksheet.cell(row=i + 1, column=j + 1).value = cell.value
                            _worksheet.cell(row=i + 1, column=j + 1).style = cell.style

                            if i == _workbook[_sheet_name].max_row - 1:
                                _worksheet.column_dimensions[get_column_letter(j + 1)].width = _workbook[_sheet_name].column_dimensions[get_column_letter(j + 1)].width

                        # self.workbook.move_sheet(_worksheet, offset=-len(self.workbook.sheetnames) + 1)

            _workbook.close()

        if os.path.dirname(filename) != "" and not os.path.exists(os.path.dirname(filename)):
            os.makedirs(os.path.dirname(filename), exist_ok=True)

        self.workbook.save(filename)

        if close:
            self.workbook.close()


def dataframe2excel(data, excel_writer, sheet_name=None, title=None, header=True, theme_color="2639E9", condition_color=None, fill=True, percent_cols=None, condition_cols=None, custom_cols=None, custom_format="#,##0", color_cols=None, percent_rows=None, condition_rows=None, custom_rows=None, color_rows=None, start_col=2, start_row=2, mode="replace", figures=None, figsize=(600, 350), writer_params={}, **kwargs):
    """
    向excel文件中插入指定样式的dataframe数据

    :param data: 需要保存的dataframe数据，index默认不保存，如果需要保存先 .reset_index().rename(columns={"index": "索引名称"}) 再保存，有部分索引 reset_index 之后是 0 而非 index，根据实际情况进行修改
    :param excel_writer: 需要保存到的 excel 文件路径或者 ExcelWriter
    :param sheet_name: 需要插入内容的sheet，如果是 Worksheet，则直接向 Worksheet 插入数据
    :param title: 是否在dataframe之前的位置插入一个标题
    :param figures: 需要数据表与标题之间插入的图片，支持一次性传入多张图片的路径，会根据传入顺序依次插入
    :param figsize: 插入图像的大小，为了统一排版，目前仅支持设置一个图片大小，默认: (600, 350) (长度, 高度)
    :param header: 是否存储dataframe的header，暂不支持多级表头
    :param theme_color: 主题色
    :param condition_color: 条件格式主题颜色，不传默认为 theme_color
    :param fill: 是否使用单元个颜色填充样式还是使用边框样式
    :param percent_cols: 需要显示为百分数的列，仅修改显示格式，不更改数值
    :param condition_cols: 需要显示条件格式的列（无边框渐变数据条）
    :param color_cols: 需要显示为条件格式颜色填充的列（单元格填充渐变色）
    :param custom_cols: 需要显示自定义格式的列，与 custom_format 参数搭配使用
    :param custom_format: 显示的自定义格式，与 custom_cols 参数搭配使用，默认 #,##0 ，即显示为有分隔符的整数
    :param start_col: 在excel中的开始列数，默认 2，即第二列开始
    :param start_row: 在excel中的开始行数，默认 2，即第二行开始，如果 title 有值的话，会从 start_row + 2 行开始插入dataframe数据
    :param mode: excel写入的模式，可选 append 和 replace ，默认 replace ，选择 append 时会在已有的excel文件中增加内容，不覆盖原有内容
    :param writer_params: 透传至 ExcelWriter 内的参数
    :param **kwargs: 其他参数，透传至 insert_df2sheet 方法，例如 传入 auto_width=True 会根据内容自动调整列宽

    :return: 返回插入元素最后一列之后、最后一行之后的位置

    **参考样例**

    >>> writer = ExcelWriter(theme_color='3f1dba')
    >>> worksheet = writer.get_sheet_by_name("模型报告")
    >>> end_row, end_col = writer.insert_value2sheet(worksheet, "B2", value="模型报告", style="header")
    >>> end_row, end_col = writer.insert_value2sheet(worksheet, "B4", value="模型报告", style="header", end_space="D4")
    >>> end_row, end_col = writer.insert_value2sheet(worksheet, "B6", value="当前模型主要为评分卡模型", style="header_middle", auto_width=True)
    >>> # 单层索引保存样例
    >>> sample = pd.DataFrame(np.concatenate([np.random.random_sample((10, 10)) * 40, np.random.randint(0, 3, (10, 2))], axis=1), columns=[f"B{i}" for i in range(10)] + ["target", "type"])
    >>> end_row, end_col = writer.insert_df2sheet(worksheet, sample, (end_row + 2, column_index_from_string("B")))
    >>> end_row, end_col = writer.insert_df2sheet(worksheet, sample, (end_row + 2, column_index_from_string("B")), fill=True)
    >>> end_row, end_col = writer.insert_df2sheet(worksheet, sample, (end_row + 2, column_index_from_string("B")), fill=True, header=False, index=True)
    >>> end_row, end_col = writer.insert_df2sheet(worksheet, sample, (end_row + 2, column_index_from_string("B")), merge_column="target")
    >>> end_row, end_col = writer.insert_df2sheet(worksheet, sample.set_index("type"), (end_row + 2, column_index_from_string("B")), merge_column="target", index=True, fill=True)
    >>> end_row, end_col = writer.insert_df2sheet(worksheet, sample, (end_row + 2, column_index_from_string("B")), merge_column=["target", "type"])
    >>> end_row, end_col = writer.insert_df2sheet(worksheet, sample, (end_row + 2, column_index_from_string("B")), merge_column=[10, 11])
    >>> end_row, end_col = dataframe2excel(sample, writer, theme_color='3f1dba', sheet_name="模型报告", start_row=end_row + 2, percent_cols=["B2", "B6"], condition_cols=["B3", "B9"], color_cols=["B4"])
    >>> end_row, end_col = dataframe2excel(sample, writer, theme_color='3f1dba', sheet_name="模型报告", start_row=end_row + 2, percent_cols=["B2", "B6"], condition_cols=["B3", "B9"], color_cols=["B4"], title="测试样例")
    >>> end_row, end_col = dataframe2excel(sample, writer, theme_color='3f1dba', sheet_name="模型报告", start_row=end_row + 2, percent_cols=["B2", "B6"], condition_cols=["B3", "B9"], color_cols=["B4"], title="测试样例", figures=["../examples/model_report/auto_report_corr_plot.png"])
    >>> # 多层索引保存样例
    >>> multi_sample = pd.DataFrame(np.random.randint(0, 150, size=(8, 12)), columns=pd.MultiIndex.from_product([['模拟考', '正式考'], ['数学', '语文', '英语', '物理', '化学', '生物']]), index=pd.MultiIndex.from_product([['期中', '期末'], ['雷军', '李斌'], ['测试一', '测试二']]))
    >>> multi_sample.index.names = ["考试类型", "姓名", "测试"]
    >>> end_row, end_col = dataframe2excel(multi_sample, writer, theme_color='3f1dba', sheet_name="模型报告", start_row=end_row + 2, title="测试样例", index=True, header=False)
    >>> end_row, end_col = dataframe2excel(multi_sample, writer, theme_color='3f1dba', sheet_name="模型报告", start_row=end_row + 2, title="测试样例", index=True)
    >>> end_row, end_col = dataframe2excel(multi_sample, writer, theme_color='3f1dba', sheet_name="模型报告", start_row=end_row + 2, title="测试样例", index=True, fill=False)
    >>> end_row, end_col = dataframe2excel(multi_sample.reset_index(names=multi_sample.index.names, col_level=-1), writer, theme_color='3f1dba', sheet_name="模型报告", start_row=end_row + 2, title="测试样例", index=False, fill=False, merge_column=[('', '考试类型'), ('', '姓名')])
    >>> end_row, end_col = dataframe2excel(multi_sample.reset_index(names=multi_sample.index.names, col_level=-1), writer, theme_color='3f1dba', sheet_name="模型报告", start_row=end_row + 2, title="测试样例", index=False, fill=False, merge_column=[('', '考试类型')], merge=True)
    >>> end_row, end_col = dataframe2excel(multi_sample.reset_index(names=multi_sample.index.names, col_level=-1), writer, theme_color='3f1dba', sheet_name="模型报告", start_row=end_row + 2, title="测试样例", index=True, fill=True, merge_column=[('', '考试类型')], merge=True)
    >>> writer.save("./测试样例.xlsx")
    """
    if isinstance(excel_writer, ExcelWriter):
        writer = excel_writer
    else:
        writer = ExcelWriter(theme_color=theme_color, mode=mode, **writer_params)

        # if os.path.exists(excel_writer) and mode == "append":
        #     workbook = load_workbook(excel_writer)

        #     for _sheet_name in workbook.sheetnames:
        #         _worksheet = writer.get_sheet_by_name(_sheet_name or "Sheet1")

        #         for i, row in enumerate(workbook[_sheet_name].iter_rows()):
        #             for j, cell in enumerate(row):
        #                 _worksheet.cell(row=i + 1, column=j + 1, value=cell.value)

        #     workbook.close()

    if isinstance(sheet_name, Worksheet):
        worksheet = sheet_name
    else:
        worksheet = writer.get_sheet_by_name(sheet_name or "Sheet1")

    if title:
        col_width = len(data.columns) + data.index.nlevels if kwargs.get("index", False) else len(data.columns)
        start_row, end_col = writer.insert_value2sheet(worksheet, (start_row, start_col), value=title, style="header", end_space=(start_row, start_col + col_width - 1))
        start_row += 1

    if figures is not None:
        if isinstance(figures, str):
            figures = [figures]

        pic_row = start_row
        for i, pic in enumerate(figures):
            if i == 0:
                start_row, end_col = writer.insert_pic2sheet(worksheet, pic, (pic_row, start_col), figsize=figsize)
            else:
                start_row, end_col = writer.insert_pic2sheet(worksheet, pic, (pic_row, end_col - 1), figsize=figsize)

    if "merge_column" in kwargs and kwargs["merge_column"]:
        if not isinstance(kwargs["merge_column"][0], (tuple, list)):
            kwargs["merge_column"] = [c for c in data.columns if (isinstance(c, tuple) and c[-1] in kwargs["merge_column"]) or (not isinstance(c, tuple) and c in kwargs["merge_column"])]

    end_row, end_col = writer.insert_df2sheet(worksheet, data, (start_row, start_col), fill=fill, header=header, **kwargs)

    if percent_cols:
        if not isinstance(percent_cols[0], (tuple, list)):
            percent_cols = [c for c in data.columns if (isinstance(c, tuple) and c[-1] in percent_cols) or (not isinstance(c, tuple) and c in percent_cols)]
        for c in [c for c in percent_cols if c in data.columns]:
            conditional_column = get_column_letter(start_col + data.columns.get_loc(c) + data.index.nlevels if kwargs.get("index", False) else start_col + data.columns.get_loc(c))
            writer.set_number_format(worksheet, f"{conditional_column}{end_row - len(data)}:{conditional_column}{end_row - 1}", "0.00%")

    if custom_cols:
        if not isinstance(custom_cols[0], (tuple, list)):
            custom_cols = [c for c in data.columns if (isinstance(c, tuple) and c[-1] in custom_cols) or (not isinstance(c, tuple) and c in custom_cols)]
        for c in [c for c in custom_cols if c in data.columns]:
            conditional_column = get_column_letter(start_col + data.columns.get_loc(c) + data.index.nlevels if kwargs.get("index", False) else start_col + data.columns.get_loc(c))
            writer.set_number_format(worksheet, f"{conditional_column}{end_row - len(data)}:{conditional_column}{end_row - 1}", custom_format)

    if condition_cols:
        if not isinstance(condition_cols[0], (tuple, list)):
            condition_cols = [c for c in data.columns if (isinstance(c, tuple) and c[-1] in condition_cols) or (not isinstance(c, tuple) and c in condition_cols)]
        for c in [c for c in condition_cols if c in data.columns]:
            conditional_column = get_column_letter(start_col + data.columns.get_loc(c) + data.index.nlevels if kwargs.get("index", False) else start_col + data.columns.get_loc(c))
            writer.add_conditional_formatting(worksheet, f'{conditional_column}{end_row - len(data)}', f'{conditional_column}{end_row - 1}')

    if color_cols:
        if not isinstance(color_cols[0], (tuple, list)):
            color_cols = [c for c in data.columns if (isinstance(c, tuple) and c[-1] in color_cols) or (not isinstance(c, tuple) and c in color_cols)]
        for c in [c for c in color_cols if c in data.columns]:
            try:
                rule = ColorScaleRule(start_type='num', start_value=data[c].min(), start_color=condition_color or theme_color, mid_type='num', mid_value=0., mid_color='FFFFFF', end_type='num', end_value=data[c].max(), end_color=condition_color or theme_color)
                conditional_column = get_column_letter(start_col + data.columns.get_loc(c) + data.index.nlevels if kwargs.get("index", False) else start_col + data.columns.get_loc(c))
                worksheet.conditional_formatting.add(f"{conditional_column}{end_row - len(data)}:{conditional_column}{end_row - 1}", rule)
            except:
                import traceback
                traceback.print_exc()

    if percent_rows:
        if not isinstance(percent_rows[0], (tuple, list)):
            percent_rows = [c for c in data.index if (isinstance(c, tuple) and c[-1] in percent_rows) or (not isinstance(c, tuple) and c in percent_rows)]
        for c in [c for c in percent_rows if c in data.index]:
            insert_row = data.index.get_loc(c).start if data.index.nlevels > 1 and not isinstance(data.index.get_loc(c), (int, float)) else data.index.get_loc(c)
            index_row = start_row + insert_row + data.columns.nlevels if kwargs.get("header", True) else start_row + insert_row
            index_col = start_col + data.index.nlevels if kwargs.get("index", False) else start_col
            writer.set_number_format(worksheet, f"{get_column_letter(index_col)}{index_row}:{get_column_letter(index_col + len(data.columns))}{index_row}", "0.00%")

    if custom_rows:
        if not isinstance(custom_rows[0], (tuple, list)):
            custom_rows = [c for c in data.index if (isinstance(c, tuple) and c[-1] in custom_rows) or (not isinstance(c, tuple) and c in custom_rows)]
        for c in [c for c in custom_rows if c in data.index]:
            insert_row = data.index.get_loc(c).start if data.index.nlevels > 1 and not isinstance(data.index.get_loc(c), (int, float)) else data.index.get_loc(c)
            index_row = start_row + insert_row + data.columns.nlevels if kwargs.get("header", True) else start_row + insert_row
            index_col = start_col + data.index.nlevels if kwargs.get("index", False) else start_col
            writer.set_number_format(worksheet, f"{get_column_letter(index_col)}{index_row}:{get_column_letter(index_col + len(data.columns))}{index_row}", custom_format)

    if condition_rows:
        if not isinstance(condition_rows[0], (tuple, list)):
            condition_rows = [c for c in data.index if (isinstance(c, tuple) and c[-1] in condition_rows) or (not isinstance(c, tuple) and c in condition_rows)]
        for c in [c for c in condition_rows if c in data.index]:
            insert_row = data.index.get_loc(c).start if data.index.nlevels > 1 and not isinstance(data.index.get_loc(c), (int, float)) else data.index.get_loc(c)
            index_row = start_row + insert_row + data.columns.nlevels if kwargs.get("header", True) else start_row + insert_row
            index_col = start_col + data.index.nlevels if kwargs.get("index", False) else start_col
            writer.add_conditional_formatting(worksheet, f'{get_column_letter(index_col)}{index_row}', f'{get_column_letter(index_col + len(data.columns))}{index_row}')

    if color_rows:
        if not isinstance(color_rows[0], (tuple, list)):
            color_rows = [c for c in data.index if (isinstance(c, tuple) and c[-1] in color_rows) or (not isinstance(c, tuple) and c in color_rows)]
        for c in [c for c in color_rows if c in data.index]:
            try:
                insert_row = data.index.get_loc(c).start if data.index.nlevels > 1 and not isinstance(data.index.get_loc(c), (int, float)) else data.index.get_loc(c)
                rule = ColorScaleRule(start_type='num', start_value=data.loc[c].min(), start_color=condition_color or theme_color, mid_type='num', mid_value=0., mid_color='FFFFFF', end_type='num', end_value=data.loc[c].max(), end_color=condition_color or theme_color)
                index_row = start_row + insert_row + data.columns.nlevels if kwargs.get("header", True) else start_row + insert_row
                index_col = start_col + data.index.nlevels if kwargs.get("index", False) else start_col
                worksheet.conditional_formatting.add(f"{get_column_letter(index_col)}{index_row}:{get_column_letter(index_col + len(data.columns))}{index_row}", rule)
            except:
                import traceback
                traceback.print_exc()

    if not isinstance(excel_writer, ExcelWriter) and not isinstance(sheet_name, Worksheet):
        writer.save(excel_writer)

    return end_row, end_col


if __name__ == "__main__":
    end_row = 0

    writer = ExcelWriter(theme_color='3f1dba')
    worksheet = writer.get_sheet_by_name("模型报告")

    sample = pd.DataFrame(np.concatenate([np.random.random_sample((10, 10)) * 40, np.random.randint(0, 3, (10, 2))], axis=1), columns=[f"B{i}" for i in range(10)] + ["target", "type"])
    end_row, end_col = writer.insert_df2sheet(worksheet, sample, (end_row + 2, column_index_from_string("B")), fill=True, header=False, index=True)
    writer.insert_hyperlink2sheet(worksheet, "B2", hyperlink="D5")
    end_row, end_col = writer.insert_df2sheet(worksheet, sample.set_index("type"), (end_row + 2, column_index_from_string("B")), merge_column=["target"], index=True, fill=True, merge=True)
    end_row, end_col = writer.insert_df2sheet(worksheet, sample, (end_row + 2, column_index_from_string("B")), merge_column=["target", "type"], index=True, fill=False, merge=True)
    end_row, end_col = writer.insert_df2sheet(worksheet, sample.sort_values(["target", "type"]), (end_row + 2, column_index_from_string("B")), merge_column=["target"], index=True, fill=False, merge=True)
    end_row, end_col = writer.insert_df2sheet(worksheet, sample, (end_row + 2, column_index_from_string("B")))

    multi_sample = pd.DataFrame(np.random.randint(0, 150, size=(8, 12)), columns=pd.MultiIndex.from_product([['模拟考', '正式考'], ['数学', '语文', '英语', '物理', '化学', '生物']]), index=pd.MultiIndex.from_product([['期中', '期末'], ['雷军', '李斌'], ['测试一', '测试二']]))
    multi_sample.index.names = ["考试类型", "姓名", "测试"]
    end_row, end_col = dataframe2excel(multi_sample, writer, theme_color='3f1dba', sheet_name="模型报告", start_row=end_row + 2, title="测试样例", index=True, header=False, fill=False, merge_index=True)

    data = pd.read_pickle("/Users/lubberit/Downloads/black_list.pkl")
    data.index.names = ("指标名称", "命中情况")
    end_row, end_col = dataframe2excel(data, writer, sheet_name="模型报告", start_row=end_row + 2, title=None, index=True, fill=False, theme_color='3f1dba')
    data = data.reset_index(names=[("", ""), ("渠道", "时间")]).sort_values([("", ""), ("渠道", "时间")]).reset_index(drop=True)
    end_row, end_col = dataframe2excel(data, writer, sheet_name="模型报告", start_row=end_row + 2, title=None, index=False, fill=True, merge_column=[("", "")], merge=True, theme_color='3f1dba')
    end_row, end_col = dataframe2excel(data, writer, sheet_name="模型报告", start_row=end_row + 2, index=False, fill=False, merge_column=[("", "")], merge=True, auto_width=False, theme_color='3f1dba')

    for color_rows in data[data[("渠道", "时间")] == "命中率"].index:
        rule = ColorScaleRule(start_type='num', start_value=0, start_color='3f1dba', end_type='num', end_value=data.iloc[color_rows, 2:].max(), end_color='c04d9c')
        worksheet.conditional_formatting.add(f"{get_column_letter(2 + 2)}{end_row - len(data) + color_rows}:{get_column_letter(2 + len(data.columns))}{end_row - len(data) + color_rows}", rule)
        writer.set_number_format(worksheet, f"{get_column_letter(2 + 2)}{end_row - len(data) + color_rows}:{get_column_letter(2 + len(data.columns))}{end_row - len(data) + color_rows}", "0.00%")

    end_row, end_col = dataframe2excel(data.set_index([('', ''), ('渠道', '时间')]), writer, sheet_name="模型报告", condition_color="F76E6C", color_rows=["命中率"], percent_rows=["命中率"], start_row=end_row + 2, index=True, fill=False, auto_width=False, theme_color='3f1dba')

    writer.set_freeze_panes("模型报告", "B1")

    writer.save("测试样例.xlsx")
